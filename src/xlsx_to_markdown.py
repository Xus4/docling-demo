"""将 .xlsx 转为 Markdown（本地 openpyxl，不经 MinerU）；表格用 HTML 以支持合并单元格。"""

from __future__ import annotations

import html
import math
import re
from collections.abc import Callable
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


class XlsxConversionError(Exception):
    pass


def _fmt_html_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        s = "true" if value else "false"
    elif isinstance(value, float):
        if math.isfinite(value) and value == int(value):
            s = str(int(value))
        else:
            s = str(value)
    elif isinstance(value, int):
        s = str(value)
    elif isinstance(value, (datetime, date, time)):
        if isinstance(value, datetime):
            s = value.isoformat(sep=" ", timespec="seconds")
        else:
            s = value.isoformat()
    else:
        s = str(value).replace("\r\n", "\n").replace("\r", "\n")
        s = re.sub(r"\s+", " ", s).strip()
    return html.escape(s, quote=True)


def _pad_grid_rectangular(grid: list[list[Any]]) -> int:
    if not grid:
        return 0
    max_c = max(len(r) for r in grid)
    for r in grid:
        while len(r) < max_c:
            r.append(None)
    return max_c


def _grid_ncols(grid: list[list[Any]]) -> int:
    return max((len(r) for r in grid), default=0)


def _try_pop_trailing_all_skip_row(
    grid: list[list[Any]],
    skip: list[list[bool]],
    span: list[list[tuple[int, int]]],
) -> bool:
    """删除末行（该行格全部为合并覆盖），并减小向上跨越到该行的 rowspan。"""
    if not grid:
        return False
    ncols = _grid_ncols(grid)
    if ncols <= 0:
        return False
    L = len(grid) - 1
    if not all(c < len(skip[L]) and skip[L][c] for c in range(ncols)):
        return False
    for r in range(L):
        for c in range(ncols):
            if c >= len(skip[r]) or skip[r][c]:
                continue
            rs, cs = span[r][c]
            if rs > 1 and r + rs - 1 >= L:
                span[r][c] = (max(1, rs - 1), cs)
    grid.pop()
    skip.pop()
    span.pop()
    return True


def _try_pop_trailing_all_skip_col(
    grid: list[list[Any]],
    skip: list[list[bool]],
    span: list[list[tuple[int, int]]],
) -> bool:
    """删除末列（该列格全部为合并覆盖），并减小向左跨越到该列的 colspan。"""
    if not grid:
        return False
    ncols = _grid_ncols(grid)
    if ncols <= 0:
        return False
    C = ncols - 1
    if not all(
        C < len(grid[r]) and C < len(skip[r]) and skip[r][C] for r in range(len(grid))
    ):
        return False
    nrows = len(grid)
    for r in range(nrows):
        for c in range(ncols):
            if c >= len(skip[r]) or skip[r][c]:
                continue
            rs, cs = span[r][c]
            if cs > 1 and c + cs - 1 >= C:
                span[r][c] = (rs, max(1, cs - 1))
    for r in range(nrows):
        if C < len(grid[r]):
            grid[r].pop(C)
        if C < len(skip[r]):
            skip[r].pop(C)
        if C < len(span[r]):
            span[r].pop(C)
    return True


def _try_pop_trailing_neutral_row(
    grid: list[list[Any]],
    skip: list[list[bool]],
    span: list[list[tuple[int, int]]],
) -> bool:
    """末行无 skip、span 均为 (1,1)、值全空 → 整行可删（无合并依赖）。"""
    if not grid:
        return False
    ncols = _grid_ncols(grid)
    if ncols <= 0:
        return False
    L = len(grid) - 1
    if any(c >= len(skip[L]) or skip[L][c] for c in range(ncols)):
        return False
    if any(c >= len(span[L]) or span[L][c] != (1, 1) for c in range(ncols)):
        return False
    for c in range(ncols):
        v = grid[L][c] if c < len(grid[L]) else None
        if v not in (None, ""):
            return False
    grid.pop()
    skip.pop()
    span.pop()
    return True


def _try_pop_trailing_neutral_col(
    grid: list[list[Any]],
    skip: list[list[bool]],
    span: list[list[tuple[int, int]]],
) -> bool:
    """末列无 skip、span 均为 (1,1)、值全空 → 整列可删。"""
    if not grid:
        return False
    ncols = _grid_ncols(grid)
    if ncols <= 0:
        return False
    C = ncols - 1
    for r in range(len(grid)):
        if C >= len(grid[r]) or C >= len(skip[r]) or C >= len(span[r]):
            return False
        if skip[r][C]:
            return False
        if span[r][C] != (1, 1):
            return False
        v = grid[r][C]
        if v not in (None, ""):
            return False
    for r in range(len(grid)):
        grid[r].pop(C)
        skip[r].pop(C)
        span[r].pop(C)
    return True


def _shrink_grid_after_merge(
    grid: list[list[Any]],
    skip: list[list[bool]],
    span: list[list[tuple[int, int]]],
) -> None:
    """在已应用合并的前提下，安全去掉尾部「全 skip」或「全空且无合并」的行列，并修正 span。"""
    while True:
        changed = False
        while _try_pop_trailing_all_skip_row(grid, skip, span):
            changed = True
        while _try_pop_trailing_all_skip_col(grid, skip, span):
            changed = True
        while _try_pop_trailing_neutral_row(grid, skip, span):
            changed = True
        while _try_pop_trailing_neutral_col(grid, skip, span):
            changed = True
        if not changed:
            break


def _apply_merge_spans(
    sheet: Worksheet,
    grid: list[list[Any]],
    skip: list[list[bool]],
    span: list[list[tuple[int, int]]],
) -> None:
    """根据工作表合并区域设置 skip / span，并把主单元格的值写入 grid 原点。"""
    if not grid:
        return
    n_r = len(grid)
    n_c = max(len(row) for row in grid) if grid else 0
    if n_c == 0:
        return

    for mrange in sheet.merged_cells.ranges:
        r1, r2, c1, c2 = mrange.min_row, mrange.max_row, mrange.min_col, mrange.max_col
        r1i = max(r1, 1)
        r2i = min(r2, n_r)
        c1i = max(c1, 1)
        c2i = min(c2, n_c)
        if r1i > r2i or c1i > c2i:
            continue
        master_val = sheet.cell(r1, c1).value
        o_r, o_c = r1i - 1, c1i - 1
        grid[o_r][o_c] = master_val
        rs = r2i - r1i + 1
        cs = c2i - c1i + 1
        span[o_r][o_c] = (rs, cs)
        for rr in range(r1i, r2i + 1):
            for cc in range(c1i, c2i + 1):
                if rr == r1i and cc == c1i:
                    continue
                skip[rr - 1][cc - 1] = True


def _sheet_to_grid(
    sheet: Worksheet,
    *,
    max_rows: int,
    max_cols: int,
    cancel_check: Callable[[], bool] | None,
    row_cancel_stride: int,
) -> tuple[list[list[Any]], list[list[bool]], list[list[tuple[int, int]]], bool, bool]:
    """返回 (grid, skip, span, truncated_rows, truncated_cols)。表格渲染用 HTML rowspan/colspan。"""
    mr = sheet.max_row or 1
    mc = sheet.max_column or 1
    truncated_rows = mr > max_rows
    truncated_cols = mc > max_cols
    use_r = min(mr, max_rows)
    use_c = min(mc, max_cols)
    grid: list[list[Any]] = [[None for _ in range(use_c)] for _ in range(use_r)]

    stride = max(1, int(row_cancel_stride))
    for r_idx, row in enumerate(
        sheet.iter_rows(
            min_row=1,
            max_row=use_r,
            min_col=1,
            max_col=use_c,
            values_only=False,
        ),
        start=1,
    ):
        if r_idx == 1 or (r_idx % stride) == 0:
            if cancel_check is not None and cancel_check():
                raise XlsxConversionError("任务已取消")
        for cell in row:
            c0 = cell.column - 1
            r0 = cell.row - 1
            if 0 <= r0 < use_r and 0 <= c0 < use_c:
                grid[r0][c0] = cell.value

    # 先铺满矩形，再应用合并，再安全裁剪尾部（全 skip 行/列会同步减小 rowspan/colspan；
    # 全空且无合并的末行/末列直接删除）。尺寸仍由 max_rows / max_cols 约束。
    n_c = _pad_grid_rectangular(grid)
    n_r = len(grid)
    skip = [[False for _ in range(n_c)] for _ in range(n_r)]
    span: list[list[tuple[int, int]]] = [[(1, 1) for _ in range(n_c)] for _ in range(n_r)]
    _apply_merge_spans(sheet, grid, skip, span)
    _shrink_grid_after_merge(grid, skip, span)
    return grid, skip, span, truncated_rows, truncated_cols


def _grid_to_html_table(
    grid: list[list[Any]],
    skip: list[list[bool]],
    span: list[list[tuple[int, int]]],
) -> str:
    if not grid:
        return "<p><em>（本区域无单元格数据）</em></p>\n\n"
    ncols = max(len(r) for r in grid)
    lines: list[str] = [
        '<table border="1" cellpadding="4" cellspacing="0">\n',
        "<tbody>\n",
    ]
    for r in range(len(grid)):
        cells_html: list[str] = []
        for c in range(ncols):
            if c >= len(skip[r]) or skip[r][c]:
                continue
            rs, cs = span[r][c]
            attrs = []
            if rs > 1:
                attrs.append(f' rowspan="{rs}"')
            if cs > 1:
                attrs.append(f' colspan="{cs}"')
            val = grid[r][c] if c < len(grid[r]) else None
            inner = _fmt_html_cell(val)
            tag = "th" if r == 0 else "td"
            cells_html.append(f"  <{tag}{''.join(attrs)}>{inner}</{tag}>\n")
        if not cells_html:
            continue
        lines.append("<tr>\n")
        lines.extend(cells_html)
        lines.append("</tr>\n")
    lines.append("</tbody>\n</table>\n\n")
    return "".join(lines)


def _safe_sheet_heading(name: str) -> str:
    s = name.replace("\n", " ").replace("\r", "").strip() or "Sheet"
    return s


def write_xlsx_as_markdown(
    src: Path,
    dst: Path,
    *,
    max_rows_per_sheet: int,
    max_cols: int,
    skip_hidden_sheets: bool,
    row_cancel_stride: int,
    cancel_check: Callable[[], bool] | None = None,
    progress_callback: Callable[[int, int], None] | None = None,
) -> None:
    if not src.is_file():
        raise XlsxConversionError("输入文件不存在")
    if src.suffix.lower() != ".xlsx":
        raise XlsxConversionError("仅支持 .xlsx 格式（非 .xls）")

    try:
        wb = load_workbook(
            filename=str(src),
            read_only=False,
            data_only=True,
            keep_links=False,
        )
    except Exception as exc:  # noqa: BLE001
        raise XlsxConversionError(f"无法读取 Excel 文件: {exc!s}") from exc

    try:
        visible: list[tuple[str, Worksheet]] = []
        for name in wb.sheetnames:
            ws = wb[name]
            state = getattr(ws, "sheet_state", None) or "visible"
            if skip_hidden_sheets and state != "visible":
                continue
            visible.append((name, ws))

        if not visible:
            raise XlsxConversionError("没有可导出的工作表（均为隐藏或工作簿为空）")

        total_steps = len(visible) + 1
        if progress_callback is not None:
            progress_callback(0, total_steps)

        parts: list[str] = []
        parts.append(f"# {src.name}\n\n")
        parts.append(
            "_由本机将 Excel 导出为 Markdown；**表格为 HTML**（`rowspan` / `colspan` 表示合并单元格）；"
            "首行写入表头行；图表、图片、样式不保留。_\n\n"
        )

        for idx, (raw_name, ws) in enumerate(visible):
            if cancel_check is not None and cancel_check():
                raise XlsxConversionError("任务已取消")
            title = _safe_sheet_heading(raw_name)
            parts.append(f"## 工作表：{title}\n\n")
            grid, skip, span, tr, tc = _sheet_to_grid(
                ws,
                max_rows=max(1, max_rows_per_sheet),
                max_cols=max(1, max_cols),
                cancel_check=cancel_check,
                row_cancel_stride=row_cancel_stride,
            )
            notes: list[str] = []
            if tr:
                notes.append(f"已截断行数：最多保留前 {max_rows_per_sheet} 行")
            if tc:
                notes.append(f"已截断列数：最多保留前 {max_cols} 列")
            if notes:
                parts.append("> " + "；".join(notes) + "\n\n")
            parts.append(_grid_to_html_table(grid, skip, span))
            if progress_callback is not None:
                progress_callback(idx + 1, total_steps)

        dst.parent.mkdir(parents=True, exist_ok=True)
        dst.write_text("".join(parts), encoding="utf-8")
        if progress_callback is not None:
            progress_callback(total_steps, total_steps)
    finally:
        wb.close()
