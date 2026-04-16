"""本地 .xlsx → Markdown（openpyxl）。"""

from __future__ import annotations

import os
import tempfile
import unittest
from pathlib import Path
from unittest import mock

from openpyxl import Workbook

from config import AppConfig
from src.core.service import ConversionError, ConversionService
from src.xlsx_to_markdown import XlsxConversionError, write_xlsx_as_markdown


class TestWriteXlsxAsMarkdown(unittest.TestCase):
    def test_basic_sheet(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            p = Path(td) / "a.xlsx"
            wb = Workbook()
            ws = wb.active
            assert ws is not None
            ws.title = "Data"
            ws["A1"] = "Name"
            ws["B1"] = "Count"
            ws["A2"] = "x"
            ws["B2"] = 2
            wb.save(p)
            out = Path(td) / "a.md"
            write_xlsx_as_markdown(
                p,
                out,
                max_rows_per_sheet=100,
                max_cols=50,
                skip_hidden_sheets=True,
                row_cancel_stride=32,
            )
            text = out.read_text(encoding="utf-8")
            self.assertIn("# a.xlsx", text)
            self.assertIn("## 工作表：Data", text)
            self.assertIn("<table", text)
            self.assertIn(">Name</th>", text)
            self.assertIn(">Count</th>", text)
            self.assertIn(">x</td>", text)
            self.assertIn(">2</td>", text)

    def test_hidden_sheet_skipped(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            p = Path(td) / "h.xlsx"
            wb = Workbook()
            ws0 = wb.active
            assert ws0 is not None
            ws0["A1"] = "visible"
            ws1 = wb.create_sheet("Secret")
            ws1["A1"] = "hidden"
            ws1.sheet_state = "hidden"
            wb.save(p)
            out = Path(td) / "h.md"
            write_xlsx_as_markdown(
                p,
                out,
                max_rows_per_sheet=100,
                max_cols=50,
                skip_hidden_sheets=True,
                row_cancel_stride=32,
            )
            self.assertIn("visible", out.read_text(encoding="utf-8"))
            self.assertNotIn("hidden", out.read_text(encoding="utf-8"))

    def test_merged_cells(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            p = Path(td) / "m.xlsx"
            wb = Workbook()
            ws = wb.active
            assert ws is not None
            ws["A1"] = "merged"
            ws.merge_cells("A1:B2")
            wb.save(p)
            out = Path(td) / "m.md"
            write_xlsx_as_markdown(
                p,
                out,
                max_rows_per_sheet=20,
                max_cols=20,
                skip_hidden_sheets=True,
                row_cancel_stride=8,
            )
            body = out.read_text(encoding="utf-8")
            self.assertIn("merged", body)
            self.assertIn("<table", body)
            # 尾部全 skip 行/列删除后，rowspan/colspan 会减至 1，属性可省略，仅保留一格展示合并区内容
            self.assertLessEqual(body.count("<tr>"), 2)

    def test_trailing_neutral_rows_trimmed(self) -> None:
        """末部大量空行（无合并、非 skip）应被裁掉。"""
        with tempfile.TemporaryDirectory() as td:
            p = Path(td) / "blank_tail.xlsx"
            wb = Workbook()
            ws = wb.active
            assert ws is not None
            ws["A1"] = "only"
            ws.cell(22, 1, value="")
            wb.save(p)
            out = Path(td) / "out.md"
            write_xlsx_as_markdown(
                p,
                out,
                max_rows_per_sheet=100,
                max_cols=20,
                skip_hidden_sheets=True,
                row_cancel_stride=64,
            )
            body = out.read_text(encoding="utf-8")
            self.assertIn("only", body)
            self.assertLessEqual(body.count("<tr>"), 2)

    def test_truncation_note(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            p = Path(td) / "t.xlsx"
            wb = Workbook()
            ws = wb.active
            assert ws is not None
            for i in range(30):
                ws.cell(row=i + 1, column=1, value=i)
            wb.save(p)
            out = Path(td) / "t.md"
            write_xlsx_as_markdown(
                p,
                out,
                max_rows_per_sheet=5,
                max_cols=50,
                skip_hidden_sheets=True,
                row_cancel_stride=2,
            )
            self.assertIn("已截断行数", out.read_text(encoding="utf-8"))

    def test_cancel_check(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            p = Path(td) / "c.xlsx"
            wb = Workbook()
            ws = wb.active
            assert ws is not None
            for i in range(400):
                ws.cell(row=i + 1, column=1, value=i)
            wb.save(p)
            out = Path(td) / "c.md"
            n = {"k": 0}

            def cancel() -> bool:
                n["k"] += 1
                return n["k"] >= 3

            with self.assertRaises(XlsxConversionError) as ctx:
                write_xlsx_as_markdown(
                    p,
                    out,
                    max_rows_per_sheet=500,
                    max_cols=10,
                    skip_hidden_sheets=True,
                    row_cancel_stride=1,
                    cancel_check=cancel,
                )
            self.assertIn("取消", str(ctx.exception))


class TestConversionServiceXlsx(unittest.TestCase):
    def test_xlsx_skips_mineru(self) -> None:
        with mock.patch.dict(
            os.environ, {"TABLE_SEMANTIC_ENABLE": "false"}, clear=False
        ):
            cfg = AppConfig.from_env()
        with tempfile.TemporaryDirectory() as td:
            xlsx = Path(td) / "f.xlsx"
            wb = Workbook()
            ws = wb.active
            assert ws is not None
            ws["A1"] = "h"
            ws["A2"] = "v"
            wb.save(xlsx)
            dst = Path(td) / "f.md"
            svc = ConversionService(cfg)
            with mock.patch("src.core.service.run_mineru_convert") as mineru:
                svc.convert_to_markdown(str(xlsx), str(dst))
            mineru.assert_not_called()
            self.assertTrue(dst.is_file())
            out = dst.read_text(encoding="utf-8")
            self.assertIn("<table", out)
            self.assertIn(">h</th>", out)

    def test_xlsx_table_semantic_still_runs_when_enabled(self) -> None:
        extras = {
            "TABLE_SEMANTIC_ENABLE": "true",
            "TABLE_SEMANTIC_BASE_URL": "http://127.0.0.1:1/v1",
            "TABLE_SEMANTIC_MODEL": "m",
        }
        with mock.patch.dict(os.environ, extras, clear=False):
            cfg = AppConfig.from_env()
        with tempfile.TemporaryDirectory() as td:
            xlsx = Path(td) / "f.xlsx"
            wb = Workbook()
            ws = wb.active
            assert ws is not None
            ws["A1"] = "a"
            ws["A2"] = "1"
            wb.save(xlsx)
            dst = Path(td) / "f.md"
            svc = ConversionService(cfg)
            with mock.patch("src.core.service.run_mineru_convert"):
                with mock.patch("src.core.service.augment_markdown_file") as aug:
                    svc.convert_to_markdown(str(xlsx), str(dst))
            aug.assert_called_once()

    def test_pdf_invokes_mineru(self) -> None:
        with mock.patch.dict(
            os.environ, {"TABLE_SEMANTIC_ENABLE": "false"}, clear=False
        ):
            cfg = AppConfig.from_env()
        with tempfile.TemporaryDirectory() as td:
            pdf = Path(td) / "f.pdf"
            pdf.write_bytes(b"%PDF-1.4\n")
            dst = Path(td) / "f.md"
            svc = ConversionService(cfg)

            def fake_run(*, output_path: Path, **_kw: object) -> None:
                output_path.parent.mkdir(parents=True, exist_ok=True)
                output_path.write_text("|x|\n|---|\n|y|\n", encoding="utf-8")

            with mock.patch(
                "src.core.service.run_mineru_convert", side_effect=fake_run
            ) as mineru:
                svc.convert_to_markdown(str(pdf), str(dst))
            mineru.assert_called_once()
            self.assertIn("|x|", dst.read_text(encoding="utf-8"))


if __name__ == "__main__":
    unittest.main()
